from dotenv import load_dotenv
from pathlib import Path

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

import os
import json
import logging
import uuid
import re
from datetime import datetime, timezone, timedelta
from typing import List, Optional

import bcrypt
import jwt
import cloudinary
import cloudinary.uploader
from slugify import slugify
from fastapi import FastAPI, APIRouter, Request, Response, HTTPException, Depends, UploadFile, File, Form, Query
from fastapi.staticfiles import StaticFiles
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, Field, EmailStr

# ----- Constants -----
JWT_ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 60 * 24 * 7  # 7 days for admin convenience
UPLOAD_DIR = ROOT_DIR / "uploads"
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

SYSTEMS = {
    "motor": "Motor",
    "hidraulico": "Hidráulico",
    "transmision": "Transmisión",
    "frenos": "Frenos",
    "filtros": "Filtros",
}
MODELS = ["5511-5545", "5711-5745", "6711-6745", "6911-6945", "7011-7045", "7211-7245", "8011-12045"]

# ----- DB -----
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

app = FastAPI(title="Almacén Zetor API")
api_router = APIRouter(prefix="/api")

# Mount uploads
app.mount("/api/uploads", StaticFiles(directory=str(UPLOAD_DIR)), name="uploads")

# ----- Helpers -----
def get_jwt_secret() -> str:
    return os.environ["JWT_SECRET"]

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")

def verify_password(plain: str, hashed: str) -> bool:
    return bcrypt.checkpw(plain.encode("utf-8"), hashed.encode("utf-8"))

def create_access_token(user_id: str, email: str) -> str:
    payload = {
        "sub": user_id,
        "email": email,
        "exp": datetime.now(timezone.utc) + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES),
        "type": "access",
    }
    return jwt.encode(payload, get_jwt_secret(), algorithm=JWT_ALGORITHM)

def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()

def make_slug(name: str, sku: str) -> str:
    base = slugify(name)
    return f"{base}-{sku.lower()}" if base else sku.lower()

async def get_current_user(request: Request) -> dict:
    token = request.cookies.get("access_token")
    if not token:
        auth = request.headers.get("Authorization", "")
        if auth.startswith("Bearer "):
            token = auth[7:]
    if not token:
        raise HTTPException(status_code=401, detail="No autenticado")
    try:
        payload = jwt.decode(token, get_jwt_secret(), algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "access":
            raise HTTPException(status_code=401, detail="Tipo de token inválido")
        user = await db.users.find_one({"id": payload["sub"]}, {"_id": 0, "password_hash": 0})
        if not user:
            raise HTTPException(status_code=401, detail="Usuario no encontrado")
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expirado")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Token inválido")

# ----- Models -----
class LoginRequest(BaseModel):
    email: EmailStr
    password: str

class ProductBase(BaseModel):
    sku: str
    nombre: str
    sistema: str  # motor|hidraulico|transmision|frenos|filtros
    categoria_original: Optional[str] = None
    descripcion: Optional[str] = None
    observacion_tecnica: Optional[str] = None
    compatibilidad: List[str] = Field(default_factory=list)  # ["5211","6211"]
    imagen_principal: Optional[str] = None
    galeria: List[str] = Field(default_factory=list)
    disponibilidad: str = "Disponible"  # Disponible | Bajo pedido | Agotado
    destacado: bool = False
    activo: bool = True

class ProductCreate(ProductBase):
    pass

class ProductUpdate(BaseModel):
    sku: Optional[str] = None
    nombre: Optional[str] = None
    sistema: Optional[str] = None
    categoria_original: Optional[str] = None
    descripcion: Optional[str] = None
    observacion_tecnica: Optional[str] = None
    compatibilidad: Optional[List[str]] = None
    imagen_principal: Optional[str] = None
    galeria: Optional[List[str]] = None
    disponibilidad: Optional[str] = None
    destacado: Optional[bool] = None
    activo: Optional[bool] = None

class Product(ProductBase):
    id: str
    slug: str
    created_at: str
    updated_at: str

class BlogPostBase(BaseModel):
    titulo: str
    resumen: str
    contenido: str
    imagen: Optional[str] = None
    autor: str = "Equipo Zetor"
    tags: List[str] = Field(default_factory=list)
    publicado: bool = True

class BlogPost(BlogPostBase):
    id: str
    slug: str
    created_at: str
    updated_at: str

class ContactRequest(BaseModel):
    nombre: str
    telefono: str
    email: Optional[EmailStr] = None
    ciudad: Optional[str] = None
    modelo_tractor: Optional[str] = None
    mensaje: str

# ----- Auth Endpoints -----
@api_router.post("/auth/login")
async def login(payload: LoginRequest, response: Response):
    email = payload.email.lower().strip()
    user = await db.users.find_one({"email": email})
    if not user or not verify_password(payload.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Credenciales inválidas")
    token = create_access_token(user["id"], user["email"])
    response.set_cookie(
        key="access_token", value=token, httponly=True, secure=True, samesite="none",
        max_age=ACCESS_TOKEN_EXPIRE_MINUTES * 60, path="/",
    )
    return {
        "access_token": token,
        "user": {"id": user["id"], "email": user["email"], "name": user.get("name"), "role": user.get("role")},
    }

@api_router.post("/auth/logout")
async def logout(response: Response):
    response.delete_cookie("access_token", path="/")
    return {"ok": True}

@api_router.get("/auth/me")
async def me(user: dict = Depends(get_current_user)):
    return user

# ----- Public Catalog Endpoints -----
@api_router.get("/categories")
async def get_categories():
    out = []
    for slug, label in SYSTEMS.items():
        count = await db.products.count_documents({"sistema": slug, "activo": True})
        out.append({"slug": slug, "nombre": label, "count": count})
    return out

@api_router.get("/models")
async def get_models():
    out = []
    for m in MODELS:
        count = await db.products.count_documents({"compatibilidad": m, "activo": True})
        out.append({"slug": m, "nombre": f"Zetor {m}", "count": count})
    return out

@api_router.get("/products")
async def list_products(
    sistema: Optional[str] = None,
    modelo: Optional[str] = None,
    q: Optional[str] = None,
    destacado: Optional[bool] = None,
    page: int = 1,
    limit: int = 24,
):
    query = {"activo": True}
    if sistema:
        query["sistema"] = sistema
    if modelo:
        query["compatibilidad"] = modelo
    if destacado is not None:
        query["destacado"] = destacado
    if q:
        regex = re.compile(re.escape(q), re.IGNORECASE)
        query["$or"] = [{"nombre": regex}, {"sku": regex}, {"descripcion": regex}]
    total = await db.products.count_documents(query)
    skip = max(0, (page - 1) * limit)
    cursor = db.products.find(query, {"_id": 0}).sort("destacado", -1).skip(skip).limit(limit)
    items = await cursor.to_list(length=limit)
    return {"items": items, "total": total, "page": page, "limit": limit}

@api_router.get("/products/{slug}")
async def get_product(slug: str):
    p = await db.products.find_one({"slug": slug, "activo": True}, {"_id": 0})
    if not p:
        raise HTTPException(status_code=404, detail="Producto no encontrado")
    return p

# ----- Admin Product Endpoints -----
@api_router.post("/admin/products")
async def create_product(payload: ProductCreate, user: dict = Depends(get_current_user)):
    if payload.sistema not in SYSTEMS:
        raise HTTPException(status_code=400, detail="Sistema inválido")
    existing = await db.products.find_one({"sku": payload.sku})
    if existing:
        raise HTTPException(status_code=400, detail="SKU ya existe")
    doc = payload.model_dump()
    doc["id"] = str(uuid.uuid4())
    doc["slug"] = make_slug(payload.nombre, payload.sku)
    # ensure unique slug
    n = 1
    base_slug = doc["slug"]
    while await db.products.find_one({"slug": doc["slug"]}):
        n += 1
        doc["slug"] = f"{base_slug}-{n}"
    doc["created_at"] = now_iso()
    doc["updated_at"] = now_iso()
    await db.products.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.put("/admin/products/{product_id}")
async def update_product(product_id: str, payload: ProductUpdate, user: dict = Depends(get_current_user)):
    update_doc = {k: v for k, v in payload.model_dump().items() if v is not None}
    if "sistema" in update_doc and update_doc["sistema"] not in SYSTEMS:
        raise HTTPException(status_code=400, detail="Sistema inválido")
    if "nombre" in update_doc or "sku" in update_doc:
        prod = await db.products.find_one({"id": product_id})
        if not prod:
            raise HTTPException(status_code=404, detail="Producto no encontrado")
        nombre = update_doc.get("nombre", prod["nombre"])
        sku = update_doc.get("sku", prod["sku"])
        update_doc["slug"] = make_slug(nombre, sku)
    update_doc["updated_at"] = now_iso()
    result = await db.products.update_one({"id": product_id}, {"$set": update_doc})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Producto no encontrado")
    p = await db.products.find_one({"id": product_id}, {"_id": 0})
    return p

@api_router.delete("/admin/products/{product_id}")
async def delete_product(product_id: str, user: dict = Depends(get_current_user)):
    result = await db.products.delete_one({"id": product_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Producto no encontrado")
    return {"ok": True}

@api_router.get("/admin/products")
async def admin_list_products(user: dict = Depends(get_current_user), q: Optional[str] = None, limit: int = 200):
    query = {}
    if q:
        regex = re.compile(re.escape(q), re.IGNORECASE)
        query["$or"] = [{"nombre": regex}, {"sku": regex}]
    items = await db.products.find(query, {"_id": 0}).sort("created_at", -1).limit(limit).to_list(length=limit)
    return {"items": items, "total": len(items)}

# ----- Image Upload -----
@api_router.post("/admin/upload")
async def upload_image(file: UploadFile = File(...), user: dict = Depends(get_current_user)):
    cloudinary.config(
        cloud_name=os.environ.get("CLOUDINARY_CLOUD_NAME"),
        api_key=os.environ.get("CLOUDINARY_API_KEY"),
        api_secret=os.environ.get("CLOUDINARY_API_SECRET"),
        secure=True,
    )
    allowed = {".png", ".jpg", ".jpeg", ".webp", ".gif"}
    ext = Path(file.filename).suffix.lower()
    if ext not in allowed:
        raise HTTPException(status_code=400, detail="Formato no permitido")
    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Archivo mayor a 10MB")
    try:
        result = cloudinary.uploader.upload(
            content,
            folder="zetor/productos",
            resource_type="image",
        )
        return {"url": result["secure_url"], "filename": result["public_id"]}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error subiendo imagen: {str(e)}")

@api_router.post("/admin/upload")
async def upload_image(file: UploadFile = File(...), user: dict = Depends(get_current_user)):
    allowed = {".png", ".jpg", ".jpeg", ".webp", ".gif"}
    ext = Path(file.filename).suffix.lower()
    if ext not in allowed:
        raise HTTPException(status_code=400, detail="Formato no permitido")
    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Archivo mayor a 10MB")
    try:
        result = cloudinary.uploader.upload(
            content,
            folder="zetor/productos",
            resource_type="image",
        )
        return {"url": result["secure_url"], "filename": result["public_id"]}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error subiendo imagen: {str(e)}")

# ----- Blog Endpoints -----
@api_router.get("/blog/posts")
async def list_blog():
    items = await db.blog_posts.find({"publicado": True}, {"_id": 0}).sort("created_at", -1).to_list(length=200)
    return {"items": items}

@api_router.get("/blog/posts/{slug}")
async def get_blog(slug: str):
    p = await db.blog_posts.find_one({"slug": slug, "publicado": True}, {"_id": 0})
    if not p:
        raise HTTPException(status_code=404, detail="Artículo no encontrado")
    return p

@api_router.post("/admin/blog")
async def create_blog(payload: BlogPostBase, user: dict = Depends(get_current_user)):
    doc = payload.model_dump()
    doc["id"] = str(uuid.uuid4())
    doc["slug"] = slugify(payload.titulo)
    n = 1
    base = doc["slug"]
    while await db.blog_posts.find_one({"slug": doc["slug"]}):
        n += 1
        doc["slug"] = f"{base}-{n}"
    doc["created_at"] = now_iso()
    doc["updated_at"] = now_iso()
    await db.blog_posts.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.put("/admin/blog/{post_id}")
async def update_blog(post_id: str, payload: BlogPostBase, user: dict = Depends(get_current_user)):
    update = payload.model_dump()
    update["slug"] = slugify(payload.titulo)
    update["updated_at"] = now_iso()
    result = await db.blog_posts.update_one({"id": post_id}, {"$set": update})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Artículo no encontrado")
    return await db.blog_posts.find_one({"id": post_id}, {"_id": 0})

@api_router.delete("/admin/blog/{post_id}")
async def delete_blog(post_id: str, user: dict = Depends(get_current_user)):
    result = await db.blog_posts.delete_one({"id": post_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Artículo no encontrado")
    return {"ok": True}

@api_router.get("/admin/blog")
async def admin_list_blog(user: dict = Depends(get_current_user)):
    items = await db.blog_posts.find({}, {"_id": 0}).sort("created_at", -1).to_list(length=500)
    return {"items": items}

# ----- Contact / Leads -----
@api_router.post("/contact")
async def create_contact(payload: ContactRequest):
    doc = payload.model_dump()
    doc["id"] = str(uuid.uuid4())
    doc["created_at"] = now_iso()
    doc["status"] = "nuevo"
    await db.leads.insert_one(doc)
    doc.pop("_id", None)
    return {"ok": True, "lead_id": doc["id"]}

@api_router.get("/admin/leads")
async def list_leads(user: dict = Depends(get_current_user)):
    items = await db.leads.find({}, {"_id": 0}).sort("created_at", -1).to_list(length=500)
    return {"items": items}

@api_router.delete("/admin/leads/{lead_id}")
async def delete_lead(lead_id: str, user: dict = Depends(get_current_user)):
    await db.leads.delete_one({"id": lead_id})
    return {"ok": True}

# ----- Site Settings (images & content of public site) -----
DEFAULT_SETTINGS = {
    "hero_left_image": "https://almacenzetorrepuestos.com/wp-content/uploads/2026/04/Gemini_Generated_Image_n0vlzqn0vlzqn0vl-1-scaled.png",
    "hero_right_video": "https://almacenzetorrepuestos.com/wp-content/uploads/2026/04/Agent_video_Pippit_20260429224100.mp4",
    "system_image_motor": "https://images.unsplash.com/photo-1759850425725-41216a62b6e0?crop=entropy&cs=srgb&fm=jpg&q=80&w=800",
    "system_image_hidraulico": "https://images.unsplash.com/photo-1759692071969-c32285cffc40?crop=entropy&cs=srgb&fm=jpg&q=80&w=800",
    "system_image_transmision": "https://images.unsplash.com/photo-1667339240140-1aee60bea0e5?crop=entropy&cs=srgb&fm=jpg&q=80&w=800",
    "system_image_frenos": "https://images.unsplash.com/photo-1770705950498-d373e33ecb1a?crop=entropy&cs=srgb&fm=jpg&q=80&w=800",
    "system_image_filtros": "https://images.unsplash.com/photo-1776856793085-5cfc8fefb5b8?crop=entropy&cs=srgb&fm=jpg&q=80&w=800",
    "about_mechanic_image": "https://images.unsplash.com/photo-1770705950498-d373e33ecb1a?crop=entropy&cs=srgb&fm=jpg&q=80&w=1200",
    "about_tractor_image": "https://images.unsplash.com/photo-1776856793085-5cfc8fefb5b8?crop=entropy&cs=srgb&fm=jpg&q=80&w=1200",
}

async def get_settings_dict() -> dict:
    docs = await db.site_settings.find({}, {"_id": 0}).to_list(length=200)
    settings = {**DEFAULT_SETTINGS}
    for d in docs:
        settings[d["key"]] = d["value"]
    return settings

@api_router.get("/site/settings")
async def get_site_settings():
    return await get_settings_dict()

@api_router.put("/admin/site/settings")
async def update_site_settings(payload: dict, user: dict = Depends(get_current_user)):
    """Bulk-update site settings. Pass {key: value, ...}."""
    for key, value in payload.items():
        if not isinstance(key, str):
            continue
        await db.site_settings.update_one(
            {"key": key},
            {"$set": {"key": key, "value": value, "updated_at": now_iso()}},
            upsert=True,
        )
    return await get_settings_dict()

# ----- Site config (public) -----
@api_router.get("/site/config")
async def site_config():
    return {
        "company_name": os.environ.get("COMPANY_NAME", "Almacén de Repuestos Zetor"),
        "address": os.environ.get("COMPANY_ADDRESS", ""),
        "domain": os.environ.get("COMPANY_DOMAIN", ""),
        "whatsapp": os.environ.get("WHATSAPP_NUMBER", ""),
        "systems": [{"slug": k, "nombre": v} for k, v in SYSTEMS.items()],
        "models": [{"slug": m, "nombre": f"Zetor {m}"} for m in MODELS],
    }

# ----- Bulk Import XLSX -----
SYSTEM_KEYWORDS = {
    "filtros": ["filtro"],
    "hidraulico": ["hidraul", "alce", "deposito", "depósito", "direccion", "dirección"],
    "frenos": ["freno", "albesto", "balata"],
    "transmision": ["embrague", "piñon", "piñón", "satelite", "satélite", "porta", "araña",
                    "funda", "cremallera", "terminal", "disco", "caja", "transmis"],
    "motor": ["bomba", "anillo", "biela", "valvula", "válvula", "kit motor", "kit ", "radiador",
              "ventilador", "retenedor", "multiple", "múltiple", "camisa", "empaq", "tablero",
              "silla", "farola", "pistón", "piston", "motor"],
}

def infer_system(nombre: str, categoria_original: str | None = None) -> str:
    text = f"{nombre} {categoria_original or ''}".lower()
    for sys_slug, keywords in SYSTEM_KEYWORDS.items():
        for kw in keywords:
            if kw in text:
                return sys_slug
    return "motor"  # default fallback

ALL_SERIES = ["5511-5545", "5711-5745", "6711-6745", "6911-6945", "7011-7045", "7211-7245", "8011-12045"]

def infer_compatibility(nombre: str) -> list:
    n = (nombre or "").lower()
    series = set()
    if "6911" in n: series.add("6911-6945")
    if "7011" in n: series.add("7011-7045")
    if "7211" in n or "7245" in n: series.add("7211-7245")
    if "8011" in n: series.add("8011-12045")
    if "5511" in n: series.add("5511-5545")
    if "5711" in n: series.add("5711-5745")
    if "6711" in n: series.add("6711-6745")
    if "95m" in n or "95 m" in n:
        series.update({"5511-5545", "5711-5745"})
    if "102m" in n or "102 m" in n:
        series.update({"6711-6745", "6911-6945"})
    if "110 turbo" in n:
        series.add("8011-12045")
    elif "110" in n:
        series.update({"7011-7045", "7211-7245", "8011-12045"})
    if not series:
        series = set(ALL_SERIES)
    return sorted(series, key=lambda s: ALL_SERIES.index(s))

DRIVE_RE = re.compile(r"/d/([a-zA-Z0-9_-]+)")

def drive_to_embed(url: str | None) -> str | None:
    if not url:
        return None
    m = DRIVE_RE.search(url)
    if m:
        return f"https://lh3.googleusercontent.com/d/{m.group(1)}=w1200"
    if url.startswith("http"):
        return url
    return None

@api_router.post("/admin/products/import")
async def import_products_xlsx(
    file: UploadFile = File(...),
    overwrite: bool = Form(False),
    user: dict = Depends(get_current_user),
):
    """
    Bulk-import / update products from an XLSX file.
    Expected columns (case insensitive, accents tolerated): sku, nombre, categoria,
    sistema (optional), descripcion (optional), imagen (optional), compatibilidad (optional).
    """
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Sube un archivo .xlsx")
    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Archivo mayor a 10MB")

    from io import BytesIO
    from openpyxl import load_workbook

    try:
        wb = load_workbook(BytesIO(content), data_only=True, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"No se pudo leer el archivo: {e}")

    if not rows:
        raise HTTPException(status_code=400, detail="Archivo vacío")

    # Header row → normalized
    def norm(s):
        if s is None:
            return ""
        s = str(s).strip().lower()
        for a, b in [("á", "a"), ("é", "e"), ("í", "i"), ("ó", "o"), ("ú", "u"), ("ñ", "n")]:
            s = s.replace(a, b)
        return s

    header = [norm(c) for c in rows[0]]
    data_rows = rows[1:]

    def col_idx(*names):
        for n in names:
            if n in header:
                return header.index(n)
        return None

    idx_sku = col_idx("sku", "referencia", "codigo", "código")
    idx_nombre = col_idx("nombre", "producto", "descripcion del producto", "descripcion")
    idx_categoria = col_idx("categoria", "categoría", "tipo")
    idx_sistema = col_idx("sistema")
    idx_descripcion = col_idx("descripcion", "descripción", "detalle")
    idx_imagen = col_idx("imagen", "foto", "url imagen", "image url")
    idx_compat = col_idx("compatibilidad", "modelo", "modelos")

    if idx_sku is None or idx_nombre is None:
        raise HTTPException(
            status_code=400,
            detail="Faltan columnas obligatorias: 'sku' y 'nombre'. Encontradas: " + ", ".join(header),
        )

    created, updated, skipped, errors = 0, 0, 0, []

    for row_num, row in enumerate(data_rows, start=2):
        try:
            sku = str(row[idx_sku]).strip() if row[idx_sku] is not None else ""
            nombre = str(row[idx_nombre]).strip() if row[idx_nombre] is not None else ""
            if not sku or not nombre:
                skipped += 1
                continue

            categoria = (str(row[idx_categoria]).strip() if idx_categoria is not None and row[idx_categoria] is not None else None)
            sistema_raw = (str(row[idx_sistema]).strip().lower() if idx_sistema is not None and row[idx_sistema] is not None else None)
            sistema = sistema_raw if sistema_raw in SYSTEMS else infer_system(nombre, categoria)
            descripcion = (str(row[idx_descripcion]).strip() if idx_descripcion is not None and row[idx_descripcion] is not None else f"Repuesto Zetor — {nombre}. Referencia {sku}.")
            imagen_raw = (str(row[idx_imagen]).strip() if idx_imagen is not None and row[idx_imagen] is not None else None)
            imagen = drive_to_embed(imagen_raw)

            compat = infer_compatibility(nombre)
            if idx_compat is not None and row[idx_compat] is not None:
                # Allow comma- or pipe-separated overrides
                raw = str(row[idx_compat])
                manual = [c.strip() for c in re.split(r"[,;|]", raw) if c.strip()]
                if manual:
                    compat = [c for c in manual if c in ALL_SERIES] or compat

            existing = await db.products.find_one({"sku": sku})
            doc = {
                "sku": sku,
                "nombre": nombre,
                "sistema": sistema,
                "categoria_original": categoria,
                "descripcion": descripcion,
                "observacion_tecnica": "Antes de despachar validamos compatibilidad con tu modelo y número de chasis.",
                "compatibilidad": compat,
                "imagen_principal": imagen or (existing.get("imagen_principal") if existing else None),
                "galeria": existing.get("galeria", []) if existing else [],
                "disponibilidad": "Disponible",
                "destacado": existing.get("destacado", False) if existing else False,
                "activo": True,
                "updated_at": now_iso(),
            }
            if existing:
                if not overwrite:
                    # Update only when overwrite=true; otherwise skip
                    skipped += 1
                    continue
                doc["slug"] = make_slug(nombre, sku)
                # ensure unique
                n = 1
                base = doc["slug"]
                while await db.products.find_one({"slug": doc["slug"], "id": {"$ne": existing["id"]}}):
                    n += 1
                    doc["slug"] = f"{base}-{n}"
                await db.products.update_one({"id": existing["id"]}, {"$set": doc})
                updated += 1
            else:
                doc["id"] = str(uuid.uuid4())
                doc["slug"] = make_slug(nombre, sku)
                n = 1
                base = doc["slug"]
                while await db.products.find_one({"slug": doc["slug"]}):
                    n += 1
                    doc["slug"] = f"{base}-{n}"
                doc["created_at"] = now_iso()
                await db.products.insert_one(doc)
                created += 1
        except Exception as e:
            errors.append({"row": row_num, "error": str(e)})

    return {
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "errors": errors[:20],
        "total_rows": len(data_rows),
    }
@api_router.post("/blog/posts/{slug}/view")
async def register_view(slug: str):
    await db.blog_posts.update_one(
        {"slug": slug},
        {"$inc": {"views": 1}}
    )
    return {"ok": True}
    
# ----- Health -----
@api_router.get("/")
async def root():
    return {"ok": True, "service": "Zetor API"}

# ----- Admin seeding & data seeding -----
DEFAULT_BLOG_POSTS = [
    {
        "titulo": "Cómo identificar el repuesto correcto para tu Zetor 5211",
        "resumen": "Guía técnica para reconocer compatibilidad por número de chasis, modelo y sistema antes de comprar.",
        "contenido": "Identificar el repuesto correcto evita devoluciones, daños mecánicos y pérdida de tiempo en campo. En Zetor el primer paso es ubicar la placa de identificación del tractor, donde encontrarás el modelo (5211/6211/7211/8011) y el número de chasis. Con esa referencia, valida el sistema afectado: motor, hidráulico, transmisión, frenos o filtros. Cada sistema tiene sub-referencias que cambian entre años de fabricación, por eso recomendamos enviar fotos del repuesto desmontado vía WhatsApp para que nuestro equipo confirme antes del despacho.",
        "imagen": "https://images.unsplash.com/photo-1759850425725-41216a62b6e0?crop=entropy&cs=srgb&fm=jpg&q=80&w=1200",
        "autor": "Taller Zetor",
        "tags": ["compatibilidad", "5211", "guía"],
    },
    {
        "titulo": "Mantenimiento del sistema hidráulico Zetor: errores comunes",
        "resumen": "Fugas, presión baja y bomba ruidosa. Diagnóstico rápido y repuestos clave del sistema hidráulico.",
        "contenido": "El sistema hidráulico es uno de los puntos más sensibles en tractores Zetor. Una fuga lenta puede convertirse en daño costoso si no se detecta. Verifica el nivel del depósito, inspecciona retenedores y bomba de alce, y revisa las mangueras. Repuestos críticos: bomba de alce hidráulico, depósito 1 y 2 salidas, y bomba de dirección. Antes de comprar, valida la referencia con nosotros vía WhatsApp.",
        "imagen": "https://images.unsplash.com/photo-1759692071969-c32285cffc40?crop=entropy&cs=srgb&fm=jpg&q=80&w=1200",
        "autor": "Equipo Zetor",
        "tags": ["hidráulico", "mantenimiento"],
    },
    {
        "titulo": "Kit motor 102M vs 95M vs 110 Turbo: cuál necesita tu tractor",
        "resumen": "Diferencias técnicas, modelos compatibles y cuándo conviene un overhaul completo.",
        "contenido": "Los kits motor son la inversión más estratégica para extender la vida útil del tractor. El 95M es estándar para motores naturales medianos, el 102M cubre la gama media-alta, y el 110 Turbo es para potencia turboalimentada. Cada kit incluye anillos, camisas, pistones y empaquetaduras compatibles. La elección depende del modelo y del estado del bloque. Te recomendamos enviarnos el número de chasis y fotos del motor desarmado para validar la referencia exacta.",
        "imagen": "https://images.unsplash.com/photo-1667339240140-1aee60bea0e5?crop=entropy&cs=srgb&fm=jpg&q=80&w=1200",
        "autor": "Taller Zetor",
        "tags": ["motor", "kit", "overhaul"],
    },
]

async def seed_admin():
    email = os.environ["ADMIN_EMAIL"].lower().strip()
    password = os.environ["ADMIN_PASSWORD"]
    existing = await db.users.find_one({"email": email})
    if not existing:
        await db.users.insert_one({
            "id": str(uuid.uuid4()),
            "email": email,
            "password_hash": hash_password(password),
            "name": "Administrador",
            "role": "admin",
            "created_at": now_iso(),
        })
        logging.info(f"Admin seeded: {email}")
    elif not verify_password(password, existing["password_hash"]):
        await db.users.update_one(
            {"email": email},
            {"$set": {"password_hash": hash_password(password)}},
        )
        logging.info(f"Admin password updated: {email}")

async def seed_products():
    count = await db.products.count_documents({})
    if count > 0:
        return
    seed_path = ROOT_DIR.parent / "data" / "products_seed.json"
    if not seed_path.exists():
        return
    raw = json.loads(seed_path.read_text(encoding="utf-8"))
    # Compatibility heuristic for new model series
    ALL_SERIES = ["5511-5545", "5711-5745", "6711-6745", "6911-6945", "7011-7045", "7211-7245", "8011-12045"]
    for i, item in enumerate(raw):
        name_lower = item["nombre"].lower()
        compat = set()
        if "6911" in name_lower: compat.add("6911-6945")
        if "7011" in name_lower: compat.add("7011-7045")
        if "7211" in name_lower or "7245" in name_lower: compat.add("7211-7245")
        if "8011" in name_lower: compat.add("8011-12045")
        if "5511" in name_lower: compat.add("5511-5545")
        if "5711" in name_lower: compat.add("5711-5745")
        if "6711" in name_lower: compat.add("6711-6745")
        if "95m" in name_lower or "95 m" in name_lower:
            compat.update({"5511-5545", "5711-5745"})
        if "102m" in name_lower or "102 m" in name_lower:
            compat.update({"6711-6745", "6911-6945"})
        if "110 turbo" in name_lower:
            compat.add("8011-12045")
        elif "110" in name_lower:
            compat.update({"7011-7045", "7211-7245", "8011-12045"})
        if not compat:
            compat = set(ALL_SERIES)
        compat = sorted(compat, key=lambda s: ALL_SERIES.index(s))

        # Drive image link → for now use system category fallback
        imagen = None  # we won't use Drive raw URLs (not displayable). Admin can upload later.
        sys_fallbacks = {
            "motor": "https://images.unsplash.com/photo-1759850425725-41216a62b6e0?crop=entropy&cs=srgb&fm=jpg&q=80&w=800",
            "hidraulico": "https://images.unsplash.com/photo-1759692071969-c32285cffc40?crop=entropy&cs=srgb&fm=jpg&q=80&w=800",
            "transmision": "https://images.unsplash.com/photo-1667339240140-1aee60bea0e5?crop=entropy&cs=srgb&fm=jpg&q=80&w=800",
            "frenos": "https://images.unsplash.com/photo-1770705950498-d373e33ecb1a?crop=entropy&cs=srgb&fm=jpg&q=80&w=800",
            "filtros": "https://images.unsplash.com/photo-1776856793085-5cfc8fefb5b8?crop=entropy&cs=srgb&fm=jpg&q=80&w=800",
        }
        imagen = sys_fallbacks.get(item["sistema"])

        doc = {
            "id": str(uuid.uuid4()),
            "sku": item["sku"],
            "nombre": item["nombre"],
            "sistema": item["sistema"],
            "categoria_original": item.get("categoria_original"),
            "descripcion": f"Repuesto original Zetor — {item['nombre']}. Referencia {item['sku']}. Validamos compatibilidad antes del despacho.",
            "observacion_tecnica": "Antes de despachar validamos compatibilidad con tu modelo y número de chasis.",
            "compatibilidad": compat,
            "imagen_principal": imagen,
            "galeria": [],
            "disponibilidad": "Disponible",
            "destacado": i < 8,  # first 8 featured
            "activo": True,
            "slug": make_slug(item["nombre"], item["sku"]),
            "created_at": now_iso(),
            "updated_at": now_iso(),
        }
        # ensure unique slug
        n = 1
        base = doc["slug"]
        while await db.products.find_one({"slug": doc["slug"]}):
            n += 1
            doc["slug"] = f"{base}-{n}"
        await db.products.insert_one(doc)
    logging.info(f"Seeded {len(raw)} products")

async def seed_blog():
    count = await db.blog_posts.count_documents({})
    if count > 0:
        return
    for p in DEFAULT_BLOG_POSTS:
        doc = {**p}
        doc["id"] = str(uuid.uuid4())
        doc["slug"] = slugify(p["titulo"])
        doc["publicado"] = True
        doc["created_at"] = now_iso()
        doc["updated_at"] = now_iso()
        await db.blog_posts.insert_one(doc)
    logging.info("Seeded blog posts")

async def ensure_indexes():
    await db.users.create_index("email", unique=True)
    await db.products.create_index("sku", unique=True)
    await db.products.create_index("slug", unique=True)
    await db.products.create_index([("sistema", 1), ("activo", 1)])
    await db.products.create_index("compatibilidad")
    await db.blog_posts.create_index("slug", unique=True)

@app.on_event("startup")
async def on_startup():
    await ensure_indexes()
    await seed_admin()
    await seed_products()
    await seed_blog()

# Include router
app.include_router(api_router)

# CORS - allow any origin with credentials via regex (Emergent preview supports this)
app.add_middleware(
    CORSMiddleware,
    allow_origin_regex=".*",
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
